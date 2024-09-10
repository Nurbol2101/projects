import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from datetime import datetime
import os
from send2trash import send2trash

# Všechny potřebné soubory
file_path_pok = 'O:\\Work\\Resume\\07.09.2024\\Pokojská.xlsx'
file_path_naj = 'O:\\Work\\Resume\\07.09.2024\\Nájezdy+pokojů.xlsx'
file_path_pok_edited = 'O:\\Work\\Resume\\07.09.2024\\\\Pokojska (edited).xlsx'
file_path_hotova = 'O:\\Work\\Resume\\07.09.2024\\Hotova reportka.xlsx'


# Úprava excel souboru pokojská -> Zanechání pouze potřebných sloupců, jejích formatování + přídání filtrů
wb = openpyxl.load_workbook(file_path_pok)
sheet = wb['Pokojská']
columns_to_keep = ['Pokoj', 'Kapacita', 'Skupina pokojů', 'Pobyt od', 'Pobyt do', 'Osob', 'Věkové skupiny', 'Odjezd', 'Nájezd', 'Národnost']
for col in reversed(range(1, sheet.max_column + 1)):
    if sheet.cell(row=1, column=col).value not in columns_to_keep:
        sheet.delete_cols(col)
grey_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
for cell_i, cell_a in zip(sheet['H'], sheet['A']):
    if cell_i.value == "Check-in" or cell_i.value == "Check-out":
        cell_a.fill = grey_fill
sheet.insert_rows(1)

sheet.auto_filter.ref = 'A2:H2'

# Datum pro hlavní housekeeperku
current_date = datetime.now().strftime('%d.%m.%Y')
date_cell = sheet['C1']
date_cell.value = 'Datum - ' + current_date
for cell in sheet[date_cell.row]:
    cell.font = Font(size=13) 


wb.save(file_path_pok_edited)
wb.close()


#  Úprava excel souboru nájezdy pokojů -> Hledání kritéria Booking.com, vytvoření nového sloupce + zvýraznění, propojení s pomocným souborem a propsání z jednoho souboru do jíného, použítí funkce Vlookup, formatování
wb_lookup = openpyxl.load_workbook(file_path_naj)
sheet_lookup = wb_lookup['Nájezdy pokojů']
wb_results = openpyxl.load_workbook(file_path_pok_edited)
sheet_results = wb_results.active
criteria_prodejce = 'Booking.com'
sheet_results.insert_cols(1)
sheet_results['A1'] = ''

column_prodejce = 'AF'  
column_pokoj = 'C'      
column_index_prodejce = openpyxl.utils.column_index_from_string(column_prodejce) - 1
column_index_pokoj = openpyxl.utils.column_index_from_string(column_pokoj) - 1

for row in range(2, sheet_results.max_row + 1):
    room_number = sheet_results.cell(row=row, column=2).value  

    # Perform lookup
    for lookup_row in sheet_lookup.iter_rows(min_row=2, max_row=sheet_lookup.max_row, min_col=column_index_prodejce+1, max_col=column_index_prodejce+1):
        if lookup_row[0].value == criteria_prodejce and sheet_lookup.cell(row=lookup_row[0].row, column=column_index_pokoj+1).value == room_number:
            
            sheet_results.cell(row=row, column=1).value = "B"
            break  

border = Border(bottom=Side(style='thin', color='000000'))
for row in sheet_results.iter_rows(min_row=2, max_row=sheet_results.max_row):
    for cell in row:
        cell.border = border

for cell in sheet_results['A']:
    cell.font = Font(size=11, bold=True)

for row in sheet_results.iter_rows():
    sheet_results.row_dimensions[row[0].row].height = 31.5  


column_widths = {'A': 1.86, 'B': 5.14, 'C': 7.00, 'D': 12.14, 'E': 7.71, 'F': 7.71, 'G': 5.00, 'H': 11.14, 'I': 6.00, 'J': 6.80, 'K': 9.83}
for col, width in column_widths.items():
    sheet_results.column_dimensions[col].width = width


for row in sheet_results.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center')


for row in range(3, sheet_results.max_row + 1):
    cell_b = sheet_results.cell(row=row, column=2)
    if isinstance(cell_b.value, str):
        cell_b.value = cell_b.value[:3]

for row in sheet_results.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)


for row in sheet_results.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center')
for row in sheet_results.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical='top')
        
for row in sheet_results.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)



wb_results.save(file_path_hotova)


wb_lookup.close()
wb_results.close()

